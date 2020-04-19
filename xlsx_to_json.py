import openpyxl
import json

workbook = openpyxl.load_workbook('proxies.xlsx')

json_data = {}

for sheet in workbook.worksheets:
    location = sheet.title
    proxies = []
    for i in range(1, sheet.max_row + 1):
        cell_value = sheet.cell(row=i, column=1).value
        proxies.append(cell_value)
    json_data[location] = proxies

with open('proxies.json', 'w') as f:
    f.write(json.dumps(json_data, indent=2))
